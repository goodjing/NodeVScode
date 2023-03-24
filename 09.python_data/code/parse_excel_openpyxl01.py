from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# 创建excel文件
# wb = Workbook()  # 创建一个 Workbook
# ws = wb.active  # 获取被激活的 worksheet
# ws['A1'] = 'new data'  # 设置内容
# wb.save('test.xlsx')  # 按照指定路径以及文件名保存
# print(type(wb))

# 读取Excel文件
wbr = load_workbook('test.xlsx')  # 加载/读取一个已存在的Excel文件
'''
load_workbook(filename, read_only=False, keep_vba=False, data_only=False, keep_links=True)
read_only：是否只读模式，对于超大型文件（只读不写），要提升效率有帮助
keep_vba：是否保留 vba 代码，即打开 Excel 文件时，开启并保留宏
data_only：是否将公式转换为结果，即包含公式的单元格，是否显示最近的计算结果
keep_links：是否保留外部链接
'''
# print(type(wbr))
# print(wbr.sheetnames)

# 修改Excel文件
wbw = load_workbook('test.xlsx')
wsw = wbw.active
wsw['A2'] = 'data2'  # 修改A2
wsw['A3'] = 'data3'  # 修改A3
# wbw.save('test.xlsx')  # 保存修改的内容，可以保存在原文件或者另存为新的文件

# 创建sheet表
# wb = load_workbook('test.xlsx')
#
# ws1 = wb.create_sheet()  # 在末尾插入工作表
# ws2 = wb.create_sheet(index=0)  # 在开头插入工作表
# ws3 = wb.create_sheet('try', 2)  # 在位置3(下标从0开始)插入工作表’try'
# ws4 = wb.create_sheet('try2', -3)
# wb.save('test.xlsx')

# 修改sheet表名称
# wb = load_workbook('test.xlsx')
# ws = wb.active  # 获取当前工作表
# print(f'当前的工作表为：{ws.title}')
# ws.title = 'sheet two'
# print(f'修改工作表为：{ws.title}')
# wb.save('test.xlsx')

# 根据sheet名称获取sheet对象
# wb = load_workbook('test.xlsx')
# ws = wb['Sheet']
# wsv = ws['A2'].value
# print(wsv)

# 删除工作表
# wb = load_workbook('test.xlsx')
# ws = wb['try']
# wb.remove(ws)
# wb.save('test.xlsx')

# 获取/设置某个单元格
# wb = load_workbook('test.xlsx')
# ws = wb['Sheet']
# cel = ws.cell(row=3, column=1)  # row为行，column为列
# print(cel, cel.value)  # 获取单元格的值
# ws.cell(row=1, column=7, value='usee')  # 修改单元格内容
# print(ws.cell(1, 7).value)  # 查看修改后的单元格的值

# 获取某行或某列的数据
# wb = load_workbook('test.xlsx')
# ws = wb.active
# rows = ws['3']  # 定位到第三行
# print(rows)
# for rowsv in rows:  # 获取行的值
#     print(rowsv.value, end=' ')  # 输出第三行每一个cell对象
# print('\n')
# columns = ws['C']  # 定位到C列
# for columnsv in columns:  # 获取列的值
#     print(columnsv.value, end=' ')  # 输出C列每一个cell对象

# 获取多行或者多列的数据
