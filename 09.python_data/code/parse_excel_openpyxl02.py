from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment, PatternFill
from config import *


class ParseExcel(object):
    def __init__(self):
        self.wk = load_workbook(execlPath, data_only=True)
        self.execlFile = execlPath

    def getSheetByName(self, sheetName):
        """获取sheet对象"""
        sheet = self.wk[sheetName]
        return sheet

    def getRowNum(self, sheet):
        """获取有效数据的最大行号"""
        return sheet.max_row

    def getColsNum(self, sheet):
        """获取有效数据的最大列数"""
        return sheet.max_column

    def getRowValues(self, sheet, rowNum):
        """获取某一行的数据"""
        maxColsNum = self.getColsNum(sheet)
        rowValues = []
        for colsNum in range(1, maxColsNum + 1):
            value = sheet.cell(rowNum, colsNum).value
            if value is None:
                value = ''
            rowValues.append(value)
        return tuple(rowValues)

    def getColumnValues(self, sheet, columnNum):
        """获取某一列的数据"""
        maxRowNum = self.getRowNum(sheet)
        columnValues = []
        for rowNum in range(2, maxRowNum + 1):
            value = sheet.cell(rowNum, columnNum).value
            if value is None:
                value = ''
            columnValues.append(value)
        return tuple(columnValues)

    def getValueOfCell(self, sheet, rowNum, columnNum):
        """获取某一个单元格的数据"""
        value = sheet.cell(rowNum, columnNum).value
        if value is None:
            value = ''
        return value

    def getAllValuesOfSheet(self, sheet):
        """获取某一个sheet页的所有测试数据，返回一个元组组成的列表"""
        maxRowNum = self.getRowNum(sheet)
        columnNum = self.getColsNum(sheet)
        allValues = []
        for row in range(2, maxRowNum + 1):
            rowValues = []
            for column in range(1, columnNum + 1):
                value = sheet.cell(row, column).value
                if value is None:
                    value = ''
                rowValues.append(value)
            allValues.append(tuple(rowValues))
        return allValues

    def setCellValue(self, sheet, row, column, text):
        sheet.cell(row, column, value=str(text))

    def setRightCellStyle(self, sheet, row, column):
        right_12_font = Font(name='宋体', size=12, italic=False, color='000000', bold=True)  # 黑色
        sheet.cell(row, column).font = right_12_font

    def setWrongCellStyle(self, sheet, row, column):
        wrong_12_font = Font(name='宋体', size=12, italic=False, color='FF0000', bold=True)  # 红色
        sheet.cell(row, column).font = wrong_12_font

    # 保存execl数据
    def saveData(self):
        self.wk.save(execlPath)


if __name__ == '__main__':
    execl = ParseExcel()
    # sheet = execl.getSheetByName('Leave')
    # allvalue = execl.getAllValuesOfSheet(sheet)
    # print('allValue:{}'.format(allvalue))
    # corpus = execl.getRowValues(sheet, 2)
    # print(corpus)
    print(execl.getSheetByName().getRowNum())
